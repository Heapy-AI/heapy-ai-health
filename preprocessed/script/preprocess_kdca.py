#!/usr/bin/env python3
"""① 원천 KDCA(data/disease_info/*.xlsx의 API URL) → ② 전처리(preprocessed/disease_info/kdca).

각 질병 URL을 호출(레거시 SSL 우회)해 XML을 받고, 섹션(정의/원인/증상…)만 정제해 보존한다.
- 첨부파일 URL 항목 제외, HTML 태그 정리(표는 [표 생략]), 같은 섹션 연속 병합
- 청킹은 하지 않는다(크기적응형 청킹은 chunk_disease.py). 재청킹 시 여기만 다시 읽으면 됨.
- 원본 XML은 캐시 폴더에 저장해 재fetch를 피한다. 출처 URL에는 TOKEN을 넣지 않는다.
"""
from __future__ import annotations

import html
import json
import re
import ssl
import sys
import time
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
import requests
import urllib3
from requests.adapters import HTTPAdapter
from urllib3.util.ssl_ import create_urllib3_context

sys.stdout.reconfigure(encoding="utf-8")
urllib3.disable_warnings()

ROOT = Path(__file__).resolve().parents[2]   # preprocessed/script/ → repo 루트
DATA = ROOT / "data" / "disease_info"
OUT_DIR = ROOT / "preprocessed" / "disease_info" / "kdca"
XML_CACHE = ROOT / "preprocessed" / "disease_info" / "_kdca_raw_xml"   # 원본 XML 캐시(재fetch 방지)

SOURCE_LABEL = "질병관리청 국가건강정보포털"
SOURCE_URL = "https://health.kdca.go.kr/healthinfo/biz/health/gnrlzHealthInfo/gnrlzHealthInfo/gnrlzHealthInfoMain.do"
REVIEW_STATUS = "SOURCE_VERIFIED"

FILE_URL = re.compile(r"^\s*https?://\S+\s*$")
_SAFE = re.compile(r'[\\/:*?"<>|]+')


class LegacyAdapter(HTTPAdapter):
    def init_poolmanager(self, *a, **k):
        ctx = create_urllib3_context()
        ctx.options |= 0x4                 # OP_LEGACY_SERVER_CONNECT
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        k["ssl_context"] = ctx
        return super().init_poolmanager(*a, **k)


def strip_html(t: str) -> str:
    t = re.sub(r"(?is)<table.*?</table>", " [표 생략] ", t)
    t = re.sub(r"<[^>]+>", " ", t)
    t = html.unescape(t)
    return re.sub(r"[ \t]+", " ", re.sub(r"\n{3,}", "\n\n", t)).strip()


def parse_sections(xml_bytes: bytes):
    root = ET.fromstring(xml_bytes)
    merged = []
    for cl in root.iter("cntntsCl"):
        nm = (cl.findtext("CNTNTS_CL_NM") or "").strip()
        cn = (cl.findtext("CNTNTS_CL_CN") or "").strip()
        if not cn or FILE_URL.match(cn):
            continue
        cn = strip_html(cn)
        if len(cn) < 10:
            continue
        if merged and merged[-1]["name"] == nm:
            merged[-1]["content"] += "\n" + cn
        else:
            merged.append({"name": nm, "content": cn})
    return merged


def main() -> int:
    xlsx = next(DATA.glob("*국가건강정보포털*API*.xlsx"))
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    rows = [r for r in wb["Sheet1"].iter_rows(min_row=2, values_only=True) if r and r[0]]
    superclass = {r[2]: r[1] for r in wb["class"].iter_rows(min_row=2, values_only=True) if r and r[0]}
    wb.close()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    XML_CACHE.mkdir(parents=True, exist_ok=True)
    sess = requests.Session()
    sess.mount("https://", LegacyAdapter())
    sess.mount("http://", LegacyAdapter())

    ok = fail = 0
    sec_total = 0
    fails = []
    for i, r in enumerate(rows, 1):
        sn, disease, url = r[0], r[1], r[2]
        cache_fp = XML_CACHE / f"{sn}.xml"
        try:
            if cache_fp.exists():
                xb = cache_fp.read_bytes()
            else:
                for attempt in range(3):
                    try:
                        resp = sess.get(url, timeout=30, verify=False)
                        xb = resp.content
                        cache_fp.write_bytes(xb)
                        time.sleep(0.3)
                        break
                    except Exception:
                        if attempt == 2:
                            raise
                        time.sleep(1.0)
            secs = parse_sections(xb)
            if not secs:
                fail += 1
                fails.append((sn, disease, "no_sections"))
                continue
            doc = {
                "disease": disease,
                "cntnts_sn": sn,
                "superclass": superclass.get(disease),
                "source": SOURCE_URL,
                "source_label": SOURCE_LABEL,
                "review_status": REVIEW_STATUS,
                "sections": secs,
            }
            fname = _SAFE.sub("_", f"{sn}_{disease}") + ".json"
            (OUT_DIR / fname).write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
            ok += 1
            sec_total += len(secs)
        except Exception as e:
            fail += 1
            fails.append((sn, disease, str(e)[:60]))
        if i % 100 == 0:
            print(f"  진행 {i}/{len(rows)} (성공 {ok}, 실패 {fail})", flush=True)

    print(f"\n완료: 성공 {ok} / 실패 {fail} / 총 {len(rows)}  -> {OUT_DIR}")
    print(f"총 섹션 수: {sec_total} (질병당 평균 {sec_total/ok:.1f})" if ok else "")
    for sn, d, why in fails[:20]:
        print(f"  실패 sn={sn} {d}: {why}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
